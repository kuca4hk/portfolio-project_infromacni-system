import datetime

import unidecode
from django.http import JsonResponse
from django.shortcuts import render
from .models import Class, Student, StudentDetail, StudyField
from rest_framework import viewsets
from .serializers import ClassSerializer, StudentSerializer, StudentDetailSerializer, StudyFieldSerializer
from ...functions import functions as f
from ...apps.organizations.models import Organization
from ...apps.users.models import Person, PersonDetail, Address
from rest_framework.response import Response
from ...apps.users.serilazers import PersonDetailSerializer
import io
import openpyxl
from django.http import FileResponse
from rest_framework.views import APIView
# Create your views here.


class ClassViewSet(viewsets.ModelViewSet):
    queryset = Class.objects.all()
    serializer_class = ClassSerializer


class DetailClassViewSet(viewsets.ModelViewSet):
    queryset = Class.objects.all()
    serializer_class = ClassSerializer

    def retrieve(self, request, *args, **kwargs):
        class_name = kwargs.get('class_name')
        persondetail = []
        try:
            class_instance = Class.objects.get(name=class_name)
            student = StudentDetail.objects.filter(classId=class_instance.id)
            for i in student:
                st = Student.objects.get(id=i.id)
                person = Person.objects.get(id=st.personId.id)
                p = PersonDetail.objects.get(personId=person.id)
                persondetail.append(p)
            serializer = PersonDetailSerializer(persondetail, many=True)
            return Response(serializer.data)
        except Class.DoesNotExist:
            return Response({'message': 'The class does not exist'}, status=404)


class StudentDetailViewSet(viewsets.ModelViewSet):
    queryset = StudentDetail.objects.all()
    serializer_class = StudentSerializer

    def retrieve(self, request, *args, **kwargs):
        student_surname = kwargs.get('student_surname')
        student_surname_uni = unidecode.unidecode(student_surname).lower()

        try:
            persondetails = PersonDetail.objects.all()

            # Vytvoření seznamu záznamů, které odpovídají filtraci
            filtered_persondetails = [p for p in persondetails if
                                      unidecode.unidecode(p.lastName).lower() == student_surname_uni]

            # Získání odpovídajících Student objektů
            matching_students = Student.objects.filter(personId__in=[p.personId for p in filtered_persondetails])

            if matching_students.exists():
                # Získání odpovídajících StudentDetail záznamů
                student_ids = [student.id for student in matching_students]
                studendetail = StudentDetail.objects.filter(StudentId__in=student_ids)

                serializer = StudentDetailSerializer(studendetail, many=True)
                return Response(serializer.data)
            else:
                # Žádný odpovídající student nebyl nalezen
                return Response({"message": "No matching students found"}, status=404)
        except Class.DoesNotExist:
            return Response({'message': 'No matching students found'}, status=404)


class StudentDetailViewSetAll(viewsets.ModelViewSet):
    queryset = StudentDetail.objects.all()
    serializer_class = StudentDetailSerializer


class DownloadTempleta(APIView):
    def get(self, request):
        # Vytvořte XLSX soubor s falešnými daty
        wb = openpyxl.Workbook()
        ws = wb.active

        # Přidejte falešná data do buňek
        ws['A1'] = 'Příjmení'
        ws['B1'] = 'Jméno'
        ws['C1'] = 'Email zak'
        ws['D1'] = 'Email rodic 1'
        ws['E1'] = 'Email rodic 2'
        ws['F1'] = 'Telefon zak'
        ws['G1'] = 'Telefon rodic 1'
        ws['H1'] = 'Telefon rodic 2'
        ws['I1'] = ''
        ws['J1'] = 'Adresa'
        ws['K1'] = 'Město'
        ws['L1'] = 'Číslo pojišťovny'
        ws['M1'] = 'Rozhodné datum'
        ws['N1'] = 'Izoz (181105527)'
        ws['O1'] = 'CAST (01)'
        ws['P1'] = 'Rodné číslo(bez /)'
        ws['Q1'] = 'Pohlaví'
        ws['R1'] = 'Datum narození'
        ws['S1'] = 'Kvalifikátor statního občanství (Občan ČR)'
        ws['T1'] = 'Státní občanství (Česká republika)'
        ws['U1'] = 'Obec trvalého pobytu (Šestajovice)'
        ws['V1'] = 'Okres trvalého pobytu (Praha - východ)'
        ws['W1'] = 'Předchozí působiště žáka(Základní škola - z 9.ročníku)'
        ws['X1'] = 'Předchozí působiště žáka na škole (IZO školy)'
        ws['Y1'] = 'Nejvyšší dosažené vzdělání (Základní)'
        ws['Z1'] = 'Datum zahaájení studia'
        ws['AA1'] = 'Kod zahajíní studia (Přijetí do 1.ročníku)'
        ws['AB1'] = 'Datum ukončení studia (Pokud žák ukončil studium)'
        ws['AC1'] = 'Kod ukončení studia (Pokud žák ukončil studium)'
        ws['AD1'] = 'Ročník studia (Třetí ročník)'
        ws['AE1'] = 'Příznak vzdělání (Řádné vzdělávání)'
        ws['AF1'] = 'Způsob plnění docházky (Školní docházka ve škole zapsané ve školském rejstříku)'
        ws['AG1'] = 'Doba přerušení studia (kolik měsíců)'
        ws['AH1'] = 'Třída (2.alt)'
        ws['AI1'] = 'Obor (7941K41)'
        ws['AJ1'] = 'Druh vzdělání (41)'
        ws['AK1'] = 'Délka vzdělání (40)'
        ws['AL1'] = 'Forma vzdělání (Denní)'
        ws['AM1'] = 'Jazyk výuky (Český (10))'
        ws['AN1'] = 'Jazyk výuky (Anglický (02))'
        ws['AO1'] = 'Jazyk výuky (Španělský (25))'
        ws['AP1'] = ''
        ws['AQ1'] = ''
        ws['AR1'] = 'Finanocání studia (1)'
        ws['AS1'] = 'Kod zkoušky'
        ws['AT1'] = 'Kod opakování'
        ws['AU1'] = 'Kod ciziho jazyka u maturity'
        ws['AV1'] = 'Kod uspěšnosti u maturity'
        ws['AW1'] = 'Datum konání zkoušky'
        ws['AX1'] = ''
        ws['AY1'] = ''
        ws['AZ1'] = ''
        ws['BA1'] = ''
        ws['BB1'] = 'Kod změn (Beze změny)'
        ws['BC1'] = 'Datum uskutečnění změny'
        ws['BD1'] = 'Kod věty(Žák/student)'
        ws['BE1'] = 'Platnost začátku'
        ws['BF1'] = 'Platnost konce'
        ws['BG1'] = 'Anonymozovaný soubour A (ANO, jinak nic)'
        ws['BH1'] = 'KOD Zaka'
        ws['BI1'] = 'Typ třídy (Běžná třída/studijní skupina)'
        ws['BJ1'] = 'Individuální vzdělávací plán'
        ws['BK1'] = 'Rozlišení pro nadaného žáka'
        ws['BL1'] = 'Odraz odlišného kulturního prostředí nebo jiných životních podmínek žáka/žákyně do vzdělávání (dříve sociální znevýhodnění)'
        ws['BM1'] = 'Kategorie zdravotního znevýhodnění'
        ws['BN1'] = 'Identifikátor znevýhodnění'
        ws['BO1'] = 'Převažující stupeň poskytovaných opatření - vydává školské poradenské zařízení'
        ws['BP1'] = 'Prodloužená délka vzdělávání (pro ZŠ, SŠ, konzervatoře)'
        ws['BQ1'] = 'Úprava očekávaných výstupů vzdělávání (jen pro ZŠ, SŠ)'
        ws['BR1'] = 'Podpůrné opatrění (soubour B)'
        ws['BS1'] = 'RED IZO'
        ws['BT1'] = 'Oznaceni typu třídy'
        ws['BU1'] = 'IZO spz'
        ws['BV1'] = 'Datum vydaní doporučení'
        ws['BW1'] = 'Datum konce platnosti doporučení (pokud je ve 4. ročníku, předpokládá se že udělá maturitu)'
        ws['BX1'] = 'Kód normované finanční náročnosti'
        ws['BY1'] = 'Finanční náročnost'
        ws['BZ1'] = 'Datum zahájení poskytování PO'
        ws['CA1'] = 'Datum ukončení poskytování PO'
        ws['CB1'] = 'Datum skutečného zahájení poskytování'
        ws['CC1'] = 'Datum skutečného ukončení poskytování'

        # Můžete přidat další buňky a data podle potřeby
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        # Vytvořte a vrátíte odpověď s XLSX souborem
        response = FileResponse(output,
                                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="template.xlsx"'
        return response


class ExportClass(APIView):
    def get(self, request, class_name):
        response = None
        wb = openpyxl.Workbook()
        ws = wb.active

        # Nastavení záhlaví XLSX souboru
        ws.append(
            ['Jméno', 'Příjmení', 'Datum narození', 'Pohlaví', 'Státní občanství', 'Telefon zak', 'Adresa bydliště'])
        try:
            # Querry
            class_instance = Class.objects.get(name=class_name)
            students = StudentDetail.objects.filter(classId=class_instance.id)
            for i in students:
                student = Student.objects.get(id=i.id)
                persons = Person.objects.filter(id=student.personId.id).last()
                persondetail = PersonDetail.objects.get(personId=persons.id)
            # adress = Address.objects.filter(id__in=persons.values_list('birthAddressId', flat=True))

            # Přidání dat ze záznamů do souboru

                row_data = [
                        persondetail.firstName,
                        persondetail.lastName,
                        str(persondetail.birthOn),
                        str(persondetail.sex),
                        str(persondetail.citizenship),
                    ]
                ws.append(row_data)

            output = io.BytesIO()
            wb.save(output)
            output.seek(0)

            # Vytvoření a vrácení odpovědi s XLSX souborem
            response = FileResponse(output,
                                    content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            if response is not None:
                response['Content-Disposition'] = f'attachment; filename="{class_name}.xlsx"'
        except Class.DoesNotExist:
            return Response({'message': 'No matching classes found'}, status=404)

        return response


